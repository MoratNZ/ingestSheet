import openpyxl

import json

from .cellOperations import isEmptyCell, isChildMergedCell, getCellValue
from .rowOperations import parseRow
from .utility import makeCamelCase


def parseHeaders(sheet, firstHeaderRow=1, lastHeaderRow=1, labelColumn=1, maxColumnGap=1, camelCaseHeaders=False):
    """Takes an excel sheet, and returns a representation of the header rows.

    Args:
        sheet ([type]): An openpyxl worksheet object
        firstHeaderRow (int, optional): The first row that contains headers. Defaults to 1.
        lastHeaderRow (int, optional): The last row that contains headers. Defaults to the value of firstHeaderRow.
        labelColumn (int, optional): The column number that contains row labels . Defaults to 1.
        maxColumnGap (int, optional): The maximum number of blank columns allowed within the data. If more than this number of
            blank columns are encountered, data is assumed to be finished. Defaults to 1.
        camelCaseHeaders (bool, optional): Transform the headers into camelCase. Defaults to False.

    Returns:
        list of lists: one list per column, with each inner list containing the values of the header cells (with merged cells expanded)
    """

    if lastHeaderRow < firstHeaderRow:
        lastHeaderRow = firstHeaderRow

    currentColumn = labelColumn + 1
    blankColumns = 0

    returnArray = []

    for i in range(labelColumn):
        returnArray.append(None)

    while True:
        columnArray = []
        for currentRow in range(lastHeaderRow, firstHeaderRow - 1, -1):
            currentCell = sheet.cell(row=currentRow, column=currentColumn)
            if not isEmptyCell(currentCell):
                cellValue = getCellValue(currentCell)

                if camelCaseHeaders:
                    columnArray.append(makeCamelCase(cellValue))
                else:
                    columnArray.append(cellValue)

        if len(columnArray) == 0:
            blankColumns += 1

            if blankColumns > maxColumnGap:
                break

        currentColumn += 1
        returnArray.append(columnArray)

    for i in range(maxColumnGap):
        returnArray.pop()

    return returnArray


def parseSheet(sheet, firstHeaderRow=1, lastHeaderRow=1, labelColumn=1, maxColumnGap=1, camelCaseHeaders=False):
    result = {}

    headers = parseHeaders(sheet, firstHeaderRow, lastHeaderRow,
                           labelColumn, maxColumnGap, camelCaseHeaders)

    for rowIndex in range(lastHeaderRow + 1, sheet.max_row):
        labelCell = sheet.cell(row=rowIndex, column=labelColumn)

        if isEmptyCell(labelCell):
            continue

        rowLabel = labelCell.value

        rowDict = parseRow(headers, sheet[rowIndex])

        result[rowLabel] = rowDict

    return result
