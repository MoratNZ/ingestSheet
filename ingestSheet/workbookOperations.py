import openpyxl


def openWorkbook(workbookName):
    """A convenience wrapper around openpyxl's load_workbook method

    Args:
        workbookName (string): the path to the workbook you want to open

    Returns:
        openpyxl.workbook.workbook.Workbook: An openpyxl workbook object
    """
    return openpyxl.load_workbook(filename=workbookName, data_only=True)


def parseWorkbook(workbookName, sheetName=None):
    pass
